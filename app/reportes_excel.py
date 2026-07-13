"""Generación del reporte Excel respetando el formato de la plantilla."""
import io
import zipfile
import datetime
from copy import copy
from openpyxl import load_workbook


def MAY(val):
    if val is None:
        return "NO APLICA"
    s = str(val).strip()
    if not s:
        return "NO APLICA"
    return s.upper()


def posicion_fmt(val):
    # "No aplica" o vacío → "NO APLICA" (con espacio, igual que el resto de campos).
    # Cualquier otra posición → completa, en MAYÚSCULAS y sin espacios.
    if val is None or not str(val).strip() or str(val).strip().lower() == 'no aplica':
        return "NO APLICA"
    return str(val).strip().upper().replace(" ", "")


def generar_reporte(activos, ruta_plantilla, nombre_hoja):
    """Genera el Excel del reporte y devuelve un BytesIO listo para send_file.
    'activos' debe venir ya ordenado."""
    wb = load_workbook(ruta_plantilla)
    ws = wb[nombre_hoja]

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 9 and mr.max_row <= 13:
            ws.unmerge_cells(str(mr))

    estilos_ref = {}
    for col in range(1, 42):
        cell = ws.cell(row=9, column=col)
        estilos_ref[col] = {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
        }

    for row in range(9, 14):
        for col in range(1, 42):
            ws.cell(row=row, column=col).value = None

    def aplicar_estilo(row):
        for col in range(1, 42):
            cell = ws.cell(row=row, column=col)
            est = estilos_ref[col]
            cell.font = copy(est['font'])
            cell.fill = copy(est['fill'])
            cell.border = copy(est['border'])
            cell.alignment = copy(est['alignment'])
            cell.number_format = est['number_format']

    fila_inicial = 9
    total = len(activos)
    altura_fila_ref = ws.row_dimensions[9].height
    for i in range(max(total, 5)):
        row = fila_inicial + i
        aplicar_estilo(row)
        if altura_fila_ref:
            ws.row_dimensions[row].height = altura_fila_ref

    for i, activo in enumerate(activos):
        row = fila_inicial + i
        resp = activo.responsable
        det = activo.detalles

        if resp and resp.extension and resp.extension.lower() != 'no aplica':
            ws.cell(row=row, column=1).value = MAY(resp.extension)
        else:
            ws.cell(row=row, column=1).value = "NO APLICA"
        if resp and not resp.num_empleado.startswith('sin-emp-'):
            ws.cell(row=row, column=2).value = MAY(resp.num_empleado)
        else:
            ws.cell(row=row, column=2).value = "NO APLICA"
        ws.cell(row=row, column=3).value = MAY(resp.nombre) if resp else "NO APLICA"
        ws.cell(row=row, column=4).value = MAY(resp.departamento) if resp else "NO APLICA"
        ws.cell(row=row, column=5).value = MAY(resp.puesto) if resp else "NO APLICA"
        ws.cell(row=row, column=6).value = MAY(activo.num_activo)
        ws.cell(row=row, column=7).value = MAY(activo.tipo_activo)
        ws.cell(row=row, column=8).value = MAY(activo.marca)
        ws.cell(row=row, column=9).value = MAY(activo.modelo)
        ws.cell(row=row, column=10).value = MAY(activo.num_serie)
        ws.cell(row=row, column=11).value = MAY(activo.accesorio)
        ws.cell(row=row, column=12).value = MAY(det.procesador if det else None)
        ws.cell(row=row, column=13).value = MAY(det.ipv4 if det else None)
        ws.cell(row=row, column=14).value = MAY(det.mac_address if det else None)
        ws.cell(row=row, column=15).value = MAY(det.ram if det else None)
        ws.cell(row=row, column=16).value = MAY(det.tipo_almacenamiento if det else None)
        ws.cell(row=row, column=17).value = MAY(det.capacidad if det else None)
        ws.cell(row=row, column=18).value = MAY(activo.clasificacion)
        ws.cell(row=row, column=19).value = MAY(det.nombre_maquina if det else None)
        ws.cell(row=row, column=20).value = MAY(det.sistema_operativo if det else None)
        ws.cell(row=row, column=21).value = MAY(det.office if det else None)
        ws.cell(row=row, column=22).value = MAY(det.antivirus if det else None)
        ws.cell(row=row, column=23).value = MAY(det.lector_pdf if det else None)
        ws.cell(row=row, column=24).value = MAY(det.navegadores if det else None)
        ws.cell(row=row, column=25).value = MAY(det.compresor if det else None)
        ws.cell(row=row, column=26).value = MAY(det.videoconferencia if det else None)
        ws.cell(row=row, column=27).value = MAY(det.software_impresora if det else None)
        ws.cell(row=row, column=28).value = MAY(det.dominio if det else None)
        ws.cell(row=row, column=29).value = MAY(det.vpn if det else None)
        ws.cell(row=row, column=30).value = MAY(det.helpdesk_system if det else None)
        ws.cell(row=row, column=31).value = MAY(det.system_a if det else None)
        ws.cell(row=row, column=32).value = MAY(det.system_b if det else None)
        ws.cell(row=row, column=33).value = MAY(det.system_c if det else None)
        ws.cell(row=row, column=34).value = MAY(det.messaging_system if det else None)
        ws.cell(row=row, column=35).value = MAY(det.erp_system if det else None)
        ws.cell(row=row, column=36).value = MAY(det.learning_platform if det else None)
        ws.cell(row=row, column=37).value = MAY(det.quality_system if det else None)
        ws.cell(row=row, column=38).value = MAY(det.internal_database if det else None)
        ws.cell(row=row, column=39).value = MAY(det.usuario_ad if det else None)
        ws.cell(row=row, column=40).value = MAY(det.usuario_correo if det else None)
        ws.cell(row=row, column=41).value = posicion_fmt(activo.posicion)

    def clave_grupo(activo):
        if activo.responsable is None:
            return None
        if activo.responsable.num_empleado.startswith('sin-emp-'):
            return None
        return activo.responsable.id

    i = 0
    while i < total:
        clave = clave_grupo(activos[i])
        if clave is None:
            i += 1
            continue
        j = i + 1
        while j < total and clave_grupo(activos[j]) == clave:
            j += 1
        if j - i > 1:
            row_start = fila_inicial + i
            row_end = fila_inicial + j - 1
            for col in range(1, 6):
                ws.merge_cells(start_row=row_start, end_row=row_end,
                               start_column=col, end_column=col)
        i = j

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Inyectar recursos originales para preservar el logo
    preservar = {}
    archivos_a_preservar = [
        'xl/worksheets/_rels/sheet1.xml.rels',
    ]
    with zipfile.ZipFile(ruta_plantilla, 'r') as orig:
        for name in orig.namelist():
            if name.startswith('xl/drawings/') or name.startswith('xl/media/') or name in archivos_a_preservar:
                preservar[name] = orig.read(name)

    final = io.BytesIO()
    with zipfile.ZipFile(buf, 'r') as mod:
        with zipfile.ZipFile(final, 'w', zipfile.ZIP_DEFLATED) as out:
            nombres_escritos = set()
            for name in mod.namelist():
                if name in preservar:
                    out.writestr(name, preservar[name])
                else:
                    out.writestr(name, mod.read(name))
                nombres_escritos.add(name)
            for name, data in preservar.items():
                if name not in nombres_escritos:
                    out.writestr(name, data)
    final.seek(0)
    return final


def nombre_archivo():
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Inventario_Reporte_{timestamp}.xlsx"
